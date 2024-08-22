import os
import pandas as pd
import streamlit as st
from openpyxl import load_workbook,Workbook
from openpyxl.styles import Alignment

if 'data_specifics' not in st.session_state:
    list_size = 13
    st.session_state.data_specifics = {
    'Abilitati specifice': [
                    "Înțelege secvențialitatea unei sarcini",
                    "Demonstrează deprinderi de documentare",
                    "Se raportează corespunzător la cunoştințele teoretice/ aplică cunoştințele teoretice",
                    "Selectează şi utilizează instrumente adecvate (mecanice, electrice, electronice, etc.)",
                    "Face măsurători şi calcule",
                    "Notează şi înregistrează corespunzător date tehnice",
                    "Utilizează corespunzător scheme şi planuri grafice",
                    "Utilizează corespunzător tehnici şi tehnologii specifice",
                    "Descrie funcțiile şi caracteristicele de siguranță ale echipamentelor",
                    "Asamblează manual produse/ echipamente",
                    "Asamblează mechanic (electric) produse/ echipamente",
                    "Elaborează/ proiectează produse, subsisteme sau procese",
                    "Testează produse/ procese/ subsisteme",
    ],
    'Evaluare': [None for _ in range(list_size)]
}
    
if 'data_generals' not in st.session_state:
    list_size = 7
    st.session_state.data_generals = {
        'Abilitati generale' : [
            "Cunoaşte activitățile şi produsele principale ale companiei",
            "Comunică adecvat cu colegii, angajații, tutorele",
            "Comunică adecvat cu clienții (când este cazul)",
            "Demonstrează capacități de reflexie şi autoanaliză",
            "Demonstrează conduite etice",
            "Se raportează adecvat la cultura organizațională şi la contextul mai larg al întreprinderilor",
            "Analizează experiențele de învățare şi le relaționează cu oportunitățile de carieră",
        ],
        'Evaluare': [None for _ in range(list_size)]
    }
                        

data_specifics = st.session_state.data_specifics
df = pd.DataFrame(data_specifics)

data_generals = st.session_state.data_generals
dfg = pd.DataFrame(data_generals)

edited_df = st.data_editor(
    df, 
    column_config={
        'Evaluare' : st.column_config.SelectboxColumn(
            help="Calificativul oferit abilitatilor studentului",
            options=[
                "Începător",
                "Mediu",
                "Avansat"
            ],
        )
    },
    disabled=['Abilitati specifice'],
    hide_index=True
)

edited_dfg = st.data_editor(
    dfg, 
    column_config={
        'Evaluare' : st.column_config.SelectboxColumn(
            help="Calificativul oferit abilitatilor studentului",
            options=[
                "Începător",
                "Mediu",
                "Avansat"
            ],
        )
    },
    disabled=['Abilitati generale'],
    hide_index=True
)

if not edited_dfg["Evaluare"].isna().any() and not edited_df["Evaluare"].isna().any():
    if st.button("Download results"):
        file_path = 'activity_report.xlsx'
        if not os.path.exists(file_path):
        # Create a new blank workbook
            wb = Workbook()

            # Optionally, you can rename the default sheet
            ws = wb.active
            ws.title = "activity_report"

            # Save the workbook
            wb.save(file_path)
        wb = load_workbook(file_path)

        ws = wb.active
        ws['A1'].value = "ABILITĂȚI/ COMPETENȚE"
        ws['B1'].value = "Începător"
        ws['C1'].value = "Mediu"
        ws['D1'].value = "Avansat"
        ws.merge_cells('A2:D2')

        ws['A2'].value = 'Abilitati Specifice'
        index = 3

        for _,row in edited_df.iterrows():
            ws['A' + str(index)] = row["Abilitati specifice"]
            if row["Evaluare"] == "Începător":
                ws['B' + str(index)] = "X"
                ws['B' + str(index)].alignment = Alignment(horizontal='center', vertical='center')
                ws['C' + str(index)] = ""
                ws['D' + str(index)] = ""
            elif row["Evaluare"] == "Mediu":
                ws['B' + str(index)] = ""
                ws['C' + str(index)] = "X"
                ws['C' + str(index)].alignment = Alignment(horizontal='center', vertical='center')
                ws['D' + str(index)] = ""
            elif row["Evaluare"] == "Avansat":
                ws['B' + str(index)] = ""
                ws['C' + str(index)] = ""
                ws['D' + str(index)] = "X"
                ws['D' + str(index)].alignment = Alignment(horizontal='center', vertical='center')
            index+=1

        ws.merge_cells(f"A{index}:D{index}")
        ws[f"A{index}"].value = "Abilitati generale"
        index+=1
        
        for _,row in edited_dfg.iterrows():
            ws['A' + str(index)] = row["Abilitati generale"]
            if row["Evaluare"] == "Începător":
                ws['B' + str(index)] = "X"
                ws['B' + str(index)].alignment = Alignment(horizontal='center', vertical='center')
                ws['C' + str(index)] = ""
                ws['D' + str(index)] = ""
            elif row["Evaluare"] == "Mediu":
                ws['B' + str(index)] = ""
                ws['C' + str(index)] = "X"
                ws['C' + str(index)].alignment = Alignment(horizontal='center', vertical='center')
                ws['D' + str(index)] = ""
            elif row["Evaluare"] == "Avansat":
                ws['B' + str(index)] = ""
                ws['C' + str(index)] = ""
                ws['D' + str(index)] = "X"
                ws['D' + str(index)].alignment = Alignment(horizontal='center', vertical='center')
            index+=1

        try:
            wb.save(file_path)
            st.success(f"Information saved to: {file_path}")        
        except PermissionError:
            st.error("Please close the document before trying to change it")
